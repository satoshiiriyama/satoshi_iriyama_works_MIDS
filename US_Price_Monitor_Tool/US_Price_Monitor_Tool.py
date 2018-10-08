
from datetime import datetime
import datetime as dt
import time
from pytz import timezone

import pandas as pd
import numpy as np

import requests
import bs4
from urllib.request import Request, urlopen

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook


# Collecting our offcial prices

fmt_1 = "%Y-%m-%d"
fmt_2 = "%H:%M"
fmt_F = "%H%M"
fmt_W = "%y%W"

now_pst = datetime.now(timezone('US/Pacific'))
tomorrow_pst = now_pst + dt.timedelta(days=1)

tomorrow_fmt_W = tomorrow_pst.strftime(fmt_W)
today_wk = "Wk" + tomorrow_fmt_W


csv_path = "Resources/180830_price_master.csv"

price_master_df = pd.read_csv(csv_path, encoding="utf-8")

price_master_today_df = price_master_df[["Product", today_wk]]

price_master_today_df = price_master_today_df.rename(columns={today_wk: f"Official Price {today_wk}"})

price_master_today_df.to_csv("Output/price_today.csv", index=False)


# Collecting prices in Best Buy

csv_path = "Resources/180826_Best_Buy_ProductId.csv"

bby_productid_df = pd.read_csv(csv_path, encoding="utf-8")

for index, row in bby_productid_df.iterrows():
    url = row["producturl"]
    req = Request(url)
    webpage = urlopen(req).read()
    soup = bs4.BeautifulSoup(webpage, "html5lib")
    
    if str(soup).find("activated-pricing__option") >= 0:
        price = soup.select(".activated-pricing__option")
        price = price[1].getText()
        price = str(price).split(" ")[0]
        price = price.replace("$", "")
        price = float(price)
                
    else:
        price = soup.select(".priceView-hero-price.priceView-purchase-price")
        price = price[0].getText()
        price = price.replace("$", "")
        price = float(price)
        
        
    bby_productid_df.at[index, "Best Buy"] = price

bby_result_df = bby_productid_df.drop(columns=["producturl"])

bby_result_df.to_csv("Output/bby_result.csv", index=False)


# Collecting prices in B&H

csv_path = "Resources/180826_B&H_ProductId.csv"

bh_productid_df = pd.read_csv(csv_path, encoding="utf-8")

for index, row in bh_productid_df.iterrows():
    url = row["producturl"]
    
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, "html5lib")

    price = soup.select(".ypYouPay")
    price = price[0].getText().strip()
    price = price.replace("$", "")
    price = float(price)
    price

    bh_productid_df.at[index, "B&H"] = price

bh_result_df = bh_productid_df.drop(columns=["producturl"])

bh_result_df.to_csv("Output/bh_result.csv", index=False)


# Collecting prices in Amazon

csv_path = "Resources/180818_Amazon_ProductId.csv"

amazon_productid_df = pd.read_csv(csv_path, encoding="utf-8")

amazon_productid_df["asin"] = amazon_productid_df["asin"].astype(object)

amazon_asin_list = amazon_productid_df["asin"]

master_df = None

for asin in amazon_asin_list:

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
    }
   
    res = requests.get(f'https://www.amazon.com/gp/offer-listing/{asin}/ref=dp_olp_all_mbc?ie=UTF8&condition=new', headers=headers)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, "html5lib")

    
    offer_blocks = soup.select('.a-row.a-spacing-mini.olpOffer')
       
    seller_list = []
    price_list = []
        
    for offer in offer_blocks:
    
        offer_seller = offer.select('.a-spacing-none.olpSellerName')
        seller = offer_seller[0].getText().strip()
        if seller == "":
            seller = "Amazon.com"
        seller = seller + " in Amazon" 
    
        offer_price = offer.select('.a-size-large.a-color-price.olpOfferPrice.a-text-bold')
        if offer_price == []:
            price = ""
        else:
            price = offer_price[0].getText().strip()
            price = price.replace("$", "")
            price = price.replace(",", "")
            price = float(price)
        
        if seller in seller_list:
            continue
        
        else:
            seller_list.append(seller)
            price_list.append(price)
                
    product_price_df = pd.DataFrame({"Seller": seller_list,
                                         asin: price_list})

    product_price_df = product_price_df.set_index("Seller")

    master_df = pd.concat([master_df, product_price_df], axis=1, sort=False)

master_df = master_df.T
master_df = master_df.reset_index()
master_df = master_df.rename(columns={"index": "asin"})
master_df["asin"] = master_df["asin"].astype(object)
master_df = pd.merge(amazon_productid_df, master_df, on="asin", how="left")
master_df = master_df.rename(columns={"asin": "ASIN"})

amazon_result_df = master_df.drop(columns=["ASIN"])

amazon_result_df.to_csv("Output/amazon_result.csv", index=False)


# Collecting prices in Walmart

csv_path = "Resources/180818_Walmart_productId.csv"

Walmart_productid_df = pd.read_csv(csv_path, encoding="utf-8")

Walmart_productid_df["productId"] = Walmart_productid_df["productId"].astype(object)

Walmart_product_id_list = Walmart_productid_df["productId"]

seller_list = []
price_list = []
master_df = None

for product_id in Walmart_product_id_list:
    req = Request(f"https://www.walmart.com/product/{product_id}/sellers")
    webpage = urlopen(req).read()
    soup = bs4.BeautifulSoup(webpage, "html5lib")
    
    seller_list = []
    price_list = []
        
    sellers_info = soup.select('.seller-shipping-msg')

    for seller in sellers_info:
       seller = seller.getText()
       seller = seller + " in Walmart" 
       seller_list.append(seller)
  
    prices_info = soup.select('.price-characteristic')
    
    for price in prices_info:
        if str(price).find("content=") >= 0:
            price = str(price).split(" ")[2]
            price = float(price[9:-1])
            price_list.append(price)
            
    product_price_df = pd.DataFrame({"Seller": seller_list,
                                     product_id: price_list})
    product_price_df = product_price_df.set_index("Seller")
    
    master_df = pd.concat([master_df, product_price_df], axis=1, sort=False)

master_df = master_df.T
master_df = master_df.reset_index()
master_df = master_df.rename(columns={"index": "productId"})
master_df["productId"] = master_df["productId"].astype(object)
master_df = pd.merge(Walmart_productid_df, master_df, on="productId", how="left")

walmart_result_df = master_df.drop(columns=["productId"])

walmart_result_df.to_csv("Output/walmart_result.csv", index=False)


# Collecting prices in New Egg

csv_path = "Resources/180826_Newegg_ProductId.csv"

newegg_productid_df = pd.read_csv(csv_path, encoding="utf-8")

newegg_productid_df["productid"] = newegg_productid_df["productid"].astype(object)

newegg_product_id_list = newegg_productid_df["productid"]

master_df = None

for productid in newegg_product_id_list:
    
    options = Options()

    options.set_headless(True)

    driver = webdriver.Chrome(executable_path='chromedriver.exe')

    driver.get(f"https://www.newegg.com/Product/Product.aspx?Item={productid}")
    
    time.sleep(1)

    html = driver.page_source.encode('utf-8')

    soup = bs4.BeautifulSoup(html, "html.parser")
    
    seller_list = []
    price_list = []
    
    seller = soup.select(".featured-seller")[0].getText()[21:]
    if seller.find("Sold and Shipped by") > 0:
        seller = seller[:seller.find("Sold and Shipped by")]
        seller = seller + " in Newegg.com"
        seller_list.append(seller)
    else:
        seller = seller + " in Newegg.com"
        seller_list.append(seller)

        
    if soup.select(".price-current.hidden") != []:
        price = soup.select(".price-current.hidden")[0].getText()
        price = price.replace("$", "")
        price = float(price)
        price_list.append(price)
    
    elif soup.select("span.price-current") == []:
        price = ""
        price_list.append(price)    
    
    else:
        price = soup.select("span.price-current")[0].getText()
        price = price.replace("$", "")
        price = float(price)
        price_list.append(price)

    product_price_df = pd.DataFrame({"Seller": seller_list,
                                     productid: price_list})
    
    product_price_df = product_price_df.set_index("Seller")
    
    master_df = pd.concat([master_df, product_price_df], axis=1, sort=False)

master_df = master_df.T
master_df = master_df.reset_index()
master_df = master_df.rename(columns={"index": "productid"})
master_df["productid"] = master_df["productid"].astype(object)
master_df = pd.merge(newegg_productid_df, master_df, on="productid", how="left")

newegg_result_df = master_df.drop(columns=["productid"])

newegg_result_df.to_csv("Output/newegg_result.csv", index=False)


# Combine all the data tables and export to excel file and change the layout of excel file.

fmt_1 = "%Y-%m-%d"
fmt_2 = "%H:%M"
fmt_F = "%H%M"

now_pst = datetime.now(timezone('US/Pacific'))

now_date_fmt = now_pst.strftime(fmt_1)
now_time_fmt = now_pst.strftime(fmt_2)
now_time_fmt_F = now_pst.strftime(fmt_F)

price_path = "Output/price_today.csv"
bby_path = "Output/bby_result.csv"
bh_path = "Output/bh_result.csv"
amazon_path = "Output/amazon_result.csv"
walmart_path = "Output/walmart_result.csv"
newegg_path = "Output/newegg_result.csv"

price_df = pd.read_csv(price_path, encoding="utf-8")
price_df = price_df.set_index("Product")

bby_df = pd.read_csv(bby_path, encoding="utf-8")
bby_df = bby_df.set_index("Product")

bh_df = pd.read_csv(bh_path, encoding="utf-8")
bh_df = bh_df.set_index("Product")

amazon_df = pd.read_csv(amazon_path, encoding="utf-8")
amazon_df = amazon_df.set_index("Product")

walmart_df = pd.read_csv(walmart_path, encoding="utf-8")
walmart_df = walmart_df.set_index("Product")

newegg_df = pd.read_csv(newegg_path, encoding="utf-8")
newegg_df = newegg_df.set_index("Product")

master_df = None

master_df = pd.concat([master_df, price_df], axis=1, sort=False)

master_df = pd.concat([master_df, bby_df], axis=1, sort=False)

master_df = pd.concat([master_df, bh_df], axis=1, sort=False)

master_df = pd.concat([master_df, walmart_df], axis=1, sort=False)

master_df = pd.concat([master_df, newegg_df], axis=1, sort=False)

master_df = pd.concat([master_df, amazon_df], axis=1, sort=False)

master_df = master_df.reset_index()
master_df = master_df.rename(columns={"index": "Product"})

master_df.to_csv(f"Output/{now_date_fmt}_{now_time_fmt_F}_Price_Monitor.csv", index=False)
master_df.to_excel(f"Output/{now_date_fmt}_{now_time_fmt_F}_Price_Monitor.xlsx", index=False)

wb =openpyxl.load_workbook(f'Output/{now_date_fmt}_{now_time_fmt_F}_Price_Monitor.xlsx')

sheet = wb['Sheet1']

sheet.column_dimensions['A'].width = 28

sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 10
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 10
sheet.column_dimensions['H'].width = 10
sheet.column_dimensions['I'].width = 10
sheet.column_dimensions['J'].width = 10
sheet.column_dimensions['K'].width = 10
sheet.column_dimensions['L'].width = 10
sheet.column_dimensions['M'].width = 10
sheet.column_dimensions['N'].width = 10
sheet.column_dimensions['O'].width = 10
sheet.column_dimensions['P'].width = 10
sheet.column_dimensions['Q'].width = 10
sheet.column_dimensions['R'].width = 10
sheet.column_dimensions['S'].width = 10
sheet.column_dimensions['T'].width = 10
sheet.column_dimensions['U'].width = 10
sheet.column_dimensions['V'].width = 10
sheet.column_dimensions['W'].width = 10
sheet.column_dimensions['X'].width = 10
sheet.column_dimensions['Y'].width = 10
sheet.column_dimensions['Z'].width = 10
sheet.column_dimensions['AA'].width = 10
sheet.column_dimensions['AB'].width = 10
sheet.column_dimensions['AC'].width = 10
sheet.column_dimensions['AD'].width = 10
sheet.column_dimensions['AE'].width = 10
sheet.column_dimensions['AF'].width = 10
sheet.column_dimensions['AG'].width = 10
sheet.column_dimensions['AH'].width = 10
sheet.column_dimensions['AI'].width = 10
sheet.column_dimensions['AJ'].width = 10
sheet.column_dimensions['AK'].width = 10
sheet.column_dimensions['AL'].width = 10
sheet.column_dimensions['AM'].width = 10
sheet.column_dimensions['AN'].width = 10
sheet.column_dimensions['AO'].width = 10
sheet.column_dimensions['AP'].width = 10
sheet.column_dimensions['AQ'].width = 10
sheet.column_dimensions['AR'].width = 10
sheet.column_dimensions['AS'].width = 10
sheet.column_dimensions['AT'].width = 10
sheet.column_dimensions['AU'].width = 10
sheet.column_dimensions['AV'].width = 10
sheet.column_dimensions['AW'].width = 10
sheet.column_dimensions['AX'].width = 10
sheet.column_dimensions['AY'].width = 10
sheet.column_dimensions['AZ'].width = 10

sheet.freeze_panes = 'C2'

fill_1 = openpyxl.styles.PatternFill(patternType='solid', fgColor='c6e2ff', bgColor='c6e2ff')
fill_2 = openpyxl.styles.PatternFill(patternType='solid', fgColor='daedfe', bgColor='daedfe')

for i in range(1,53):
    sheet.cell(row=1, column=i).fill = fill_1
    sheet.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for x in range(2, 242):
        if x % 2 == 1:
            sheet.cell(row=x, column=i).fill = fill_2

for r in range(1, 242):
    for c in range(1, 53):
        sheet.cell(row=r, column=c).border = Border(outline=True, 
                                                    left=Side(style='dotted', color='FF000000'),
                                                    right=Side(style='dotted', color='FF000000'), 
                                                    top=Side(style='dotted', color='FF000000'), 
                                                    bottom=Side(style='dotted', color='FF000000'))
  
sheet.sheet_view.showGridLines = False

for r in range(2, 242):
    for i in range(2,53):
        if sheet.cell(row=r, column=i) != "":
            sheet.cell(row=r, column=i).number_format = '$  #,##0.00' 

for r in range(2, 242):
    for i in range(3,53):
        if sheet.cell(row=r, column=i).value == None:
            continue
        
        elif sheet.cell(row=r, column=2).value == None:
            continue
              
        elif sheet.cell(row=r, column=i).value < sheet.cell(row=r, column=2).value:
            sheet.cell(row=r, column=i).font = Font(color="FF0000", bold=True)            
            
for r in range(2, 242):
    sheet.cell(row=r, column=1).font = Font(bold=True)
    sheet.cell(row=r, column=2).font = Font(bold=True)

sheet.sheet_view.zoomScale = 85    
    
wb.save(f'Output/{now_date_fmt}_{now_time_fmt_F}_Price_Monitor.xlsx')


